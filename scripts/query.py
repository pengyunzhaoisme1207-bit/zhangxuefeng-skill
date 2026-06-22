#!/usr/bin/env python3
"""
高考志愿填报数据库查询脚本
从 Obsidian K12 资料库的 _database/ 中查询录取数据

用法:
  # 查询某省2025年录取分数
  python3 query.py --province 浙江 --mode school --year 2025

  # 查询某省某校录取分数
  python3 query.py --province 江苏 --mode school --school "南京邮电大学"

  # 按位次筛选（找位次匹配的院校）
  python3 query.py --province 广东 --mode school --rank-below 85000 --rank-above 12000

  # 查询专业录取数据
  python3 query.py --province 山东 --mode major --school "山东大学" --major "计算机"

  # 列出所有可用省份
  python3 query.py --list-provinces

  # 列出某省可用数据文件
  python3 query.py --province 浙江 --list-files

输出: Markdown 格式表格
"""

import os
import sys
import re
import argparse

# === 数据库路径映射 ===
OBSIDIAN_BASE = os.path.expanduser(
    "~/Obsidian/K12 资料库/50-高考志愿填报/_database"
)

# 省份 → 文件夹前缀映射
PROVINCE_PREFIX = {
    "河南": "01、", "湖南": "02、", "重庆": "03、", "江苏": "04、",
    "上海": "05、", "辽宁": "06、", "内蒙古": "07、", "广东": "08、",
    "浙江": "09、", "安徽": "10、", "江西": "11、", "福建": "12、",
    "河北": "13、", "山东": "14、", "天津": "15、", "湖北": "16、",
    "吉林": "17、", "北京": "18、", "广西": "19、", "贵州": "20、",
    "云南": "21、", "海南": "22、", "四川": "23、", "黑龙江": "24、",
    "陕西": "25、", "山西": "26、", "甘肃": "27、", "青海": "28、",
    "新疆": "29、", "宁夏": "30、", "西藏": "31、",
}


def find_province_dir(province):
    """根据省份名查找对应的数据库目录"""
    prefix = PROVINCE_PREFIX.get(province)
    if not prefix:
        return None
    if not os.path.isdir(OBSIDIAN_BASE):
        return None
    for d in os.listdir(OBSIDIAN_BASE):
        if os.path.isdir(os.path.join(OBSIDIAN_BASE, d)) and d.startswith(prefix):
            return os.path.join(OBSIDIAN_BASE, d)
    return None


def find_data_files(province_dir):
    """查找省份目录下的院校和专业录取数据文件"""
    school_file = None  # 院校录取分数
    major_file = None   # 专业录取分数
    plan_file = None    # 招生计划

    # 每个目录最多只扫描相关子目录
    for root, dirs, files in os.walk(province_dir):
        # Skip art exam (艺考) directories
        if '艺考' in root:
            continue

        for f in files:
            if not f.endswith('.xlsx'):
                continue
            path = os.path.join(root, f)
            basename = os.path.basename(f)

            # 院校录取分数：排除艺考相关
            if '院校录取分数' in basename and '艺术' not in basename:
                school_file = path
            # 专业录取分数
            elif '专业录取分数' in basename and '艺术' not in basename:
                if major_file is None or '20260616' in basename or '更新' in basename:
                    major_file = path
            # 招生计划
            elif '招生计划' in basename and '艺术' not in basename:
                plan_file = path

    return school_file, major_file, plan_file


def query_school_data(filepath, school=None, rank_below=None, rank_above=None, year=None, limit=50):
    """查询院校录取数据"""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return "", 0

    headers = [str(h).strip() for h in rows[0]]

    # Map column indices
    col_map = {}
    for i, h in enumerate(headers):
        for alias in ['院校名称', '学校名称']:
            if alias in h:
                col_map['school'] = i
        for alias in ['最低分位', '最低位次', '录取位次']:
            if alias in h:
                col_map['rank'] = i
        for alias in ['最低分数', '录取分数', '投档分']:
            if alias in h:
                col_map['score'] = i
        for alias in ['年份']:
            if alias == h:
                col_map['year'] = i
        for alias in ['科类']:
            if alias == h:
                col_map['category'] = i
        for alias in ['批次']:
            if alias == h:
                col_map['batch'] = i
        for alias in ['学校所在', '所在省份', '所在地']:
            if alias in h:
                col_map['location'] = i
        for alias in ['是否985', '985']:
            if alias in h:
                col_map['is_985'] = i
        for alias in ['是否211', '211']:
            if alias in h:
                col_map['is_211'] = i
        for alias in ['院校代码']:
            if alias == h:
                col_map['code'] = i

    if 'school' not in col_map or 'score' not in col_map:
        return "", 0

    # Filter rows
    results = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue

        # Year filter
        if year and col_map.get('year') is not None:
            try:
                if int(row[col_map['year']]) != year:
                    continue
            except (ValueError, TypeError):
                pass

        # School filter
        if school and col_map.get('school') is not None:
            s = str(row[col_map['school']] or '')
            if school not in s:
                continue

        # Rank filter
        if col_map.get('rank') is not None and (rank_below or rank_above):
            try:
                r = int(row[col_map['rank']])
                if rank_below and r > rank_below:
                    continue
                if rank_above and r < rank_above:
                    continue
            except (ValueError, TypeError):
                continue

        results.append(row)

    # Sort by rank ascending (use rank if available, else score)
    if col_map.get('rank') is not None:
        results.sort(key=lambda r: int(r[col_map['rank']]) if r[col_map.get('rank')] is not None else 99999999)
    elif col_map.get('score') is not None:
        results.sort(key=lambda r: int(r[col_map['score']]) if r[col_map.get('score')] is not None else 0, reverse=True)

    # Build output
    result_count = min(len(results), limit)

    # Column selection for output
    out_cols = []
    out_headers = []
    for key, label in [('year', '年份'), ('school', '院校名称'), ('category', '科类'),
                        ('batch', '批次'), ('score', '最低分'), ('rank', '最低位次'),
                        ('location', '所在地'), ('is_985', '985'), ('is_211', '211')]:
        if col_map.get(key) is not None:
            out_cols.append(col_map[key])
            out_headers.append(label)

    # Build markdown table
    md = f"| {' | '.join(out_headers)} |\n"
    md += f"| {' | '.join(['---'] * len(out_headers))} |\n"

    for row in results[:result_count]:
        vals = []
        for ci in out_cols:
            v = row[ci] if ci < len(row) else ''
            if v is None:
                v = '-'
            vals.append(str(v)[:40])
        md += f"| {' | '.join(vals)} |\n"

    wb.close()
    return md, len(results)


def list_provinces():
    """列出所有可用省份"""
    if not os.path.isdir(OBSIDIAN_BASE):
        print(f"数据库路径不存在: {OBSIDIAN_BASE}")
        return
    print(f"## 可用省份数据库\n")
    print(f"| # | 省份 | 目录 |")
    print(f"|---|------|------|")
    for num, name in sorted(PROVINCE_PREFIX.items(), key=lambda x: PROVINCE_PREFIX[x[0]]):
        d = find_province_dir(name)
        if d:
            school_f, major_f, plan_f = find_data_files(d)
            status = "✅" if school_f or major_f else "❌"
            print(f"| {PROVINCE_PREFIX[name].strip('、')} | {name} | {status} 数据就绪 |")
        else:
            print(f"| - | {name} | ❌ 未找到 |")


def list_files(province):
    """列出某省可用数据文件"""
    d = find_province_dir(province)
    if not d:
        print(f"未找到 {province} 的数据库目录")
        return

    school_f, major_f, plan_f = find_data_files(d)
    print(f"## {province} 可用数据文件\n")
    print(f"| 类型 | 文件 | 行数 |")
    print(f"|------|------|------|")

    import openpyxl
    for label, fpath in [("院校录取", school_f), ("专业录取", major_f), ("招生计划", plan_f)]:
        if fpath:
            try:
                wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                row_cnt = len(rows) - 1
                fname = os.path.basename(fpath)
                print(f"| {label} | {fname} | {row_cnt:,} 行 |")
                wb.close()
            except:
                print(f"| {label} | {os.path.basename(fpath)} | 读取失败 |")
        else:
            print(f"| {label} | - | 未找到 |")


def main():
    parser = argparse.ArgumentParser(description="高考志愿填报数据库查询")
    parser.add_argument('--province', help='省份名称，如 浙江、江苏、广东')
    parser.add_argument('--mode', choices=['school', 'major', 'plan'], default='school',
                        help='查询模式: school(院校)/major(专业)/plan(招生计划)')
    parser.add_argument('--school', help='院校名称关键词')
    parser.add_argument('--major', help='专业名称关键词（仅 major 模式）')
    parser.add_argument('--year', type=int, default=None, help='筛选年份')
    parser.add_argument('--rank-below', type=int, default=None, help='位次上限（<=此值的）')
    parser.add_argument('--rank-above', type=int, default=None, help='位次下限（>=此值的）')
    parser.add_argument('--limit', type=int, default=30, help='返回行数上限')
    parser.add_argument('--list-provinces', action='store_true', help='列出所有可用省份')
    parser.add_argument('--list-files', action='store_true', help='列出某省可用数据文件')

    args = parser.parse_args()

    if args.list_provinces:
        list_provinces()
        return

    if args.list_files or (args.province and (args.school is None and args.rank_below is None and args.rank_above is None)):
        list_files(args.province) if args.province else list_provinces()
        return

    if not args.province:
        parser.print_help()
        sys.exit(1)

    # Find province directory
    prov_dir = find_province_dir(args.province)
    if not prov_dir:
        print(f"❌ 未找到 {args.province} 的数据库目录")
        print(f"   搜索路径: {OBSIDIAN_BASE}")
        list_provinces()
        sys.exit(1)

    # Find data file
    school_f, major_f, plan_f = find_data_files(prov_dir)

    if args.mode == 'school' and school_f:
        md, total = query_school_data(
            school_f, school=args.school,
            rank_below=args.rank_below, rank_above=args.rank_above,
            year=args.year, limit=args.limit
        )
        label = f"{args.province} 院校录取数据"
    elif args.mode == 'major' and major_f:
        # For major mode, we can use the major file with school filter
        md, total = query_school_data(
            major_f, school=args.school,
            rank_below=args.rank_below, rank_above=args.rank_above,
            year=args.year, limit=args.limit
        )
        label = f"{args.province} 专业录取数据"
    else:
        print(f"❌ 未找到 {args.province} 的{args.mode}数据文件")
        sys.exit(1)

    print(f"## {label}")
    if args.school:
        print(f"> 筛选: {args.school}")
    if args.rank_below or args.rank_above:
        filters = []
        if args.rank_above: filters.append(f"位次>={args.rank_above}")
        if args.rank_below: filters.append(f"位次<={args.rank_below}")
        print(f"> 筛选: {' & '.join(filters)}")
    if args.year:
        print(f"> 年份: {args.year}")
    print()
    print(md)
    print(f"\n> 共 {total} 条结果，显示前 {min(total, args.limit)} 条")

    source = school_f or major_f or plan_f
    print(f"> 数据来源: {os.path.basename(source)}")


if __name__ == "__main__":
    main()
